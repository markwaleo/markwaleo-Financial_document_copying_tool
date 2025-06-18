import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from openpyxl import load_workbook
from datetime import datetime

log_entries = []
last_project_dir = os.getcwd()
last_source_dir = os.getcwd()

project_paths = []
source_paths = []

def add_path_row(container, path_list, path):
    row = tk.Frame(container, bg="white", height=25)
    row.pack(fill=tk.X, pady=1)
    row.grid_propagate(False)

    label = tk.Label(row, text=path, anchor="w", bg="white")
    label.grid(row=0, column=0, sticky="ew", padx=(5, 0), pady=2)

    btn = tk.Button(row, text="❌", fg="red", bg="white", bd=0, command=lambda: remove_path_row(row, path_list, path))
    btn.grid(row=0, column=1, sticky="e", padx=(5, 5), pady=2)

    row.columnconfigure(0, weight=1)
    path_list.append(path)

def remove_path_row(row, path_list, path):
    path_list.remove(path)
    row.destroy()

def select_project_dirs():
    global last_project_dir
    dir_path = filedialog.askdirectory(title="选择一个教师文件根目录", initialdir=last_project_dir)
    if dir_path and dir_path not in project_paths:
        last_project_dir = os.path.dirname(dir_path)
        add_path_row(project_scrollable_frame, project_paths, dir_path)

def select_source_dirs():
    global last_source_dir
    dir_path = filedialog.askdirectory(title="选择一个财务文件根目录", initialdir=last_source_dir)
    if dir_path and dir_path not in source_paths:
        last_source_dir = os.path.dirname(dir_path)
        add_path_row(source_scrollable_frame, source_paths, dir_path)

def clear_project_dirs():
    for widget in project_scrollable_frame.winfo_children():
        widget.destroy()
    project_paths.clear()

def clear_source_dirs():
    for widget in source_scrollable_frame.winfo_children():
        widget.destroy()
    source_paths.clear()

def extract_identifiers_from_excel(excel_path):
    try:
        wb = load_workbook(excel_path, read_only=False)
        ws = wb.active
        col_index = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value and "编号" in str(cell.value):
                    col_index = cell.column
                    break
            if col_index:
                break
        if not col_index:
            return [], "未找到编号列"
        identifiers = []
        for row in ws.iter_rows(min_row=2):
            cell = row[col_index - 1]
            if cell.value:
                identifiers.append(str(cell.value).strip())
        wb.close()
        return identifiers, None
    except Exception as e:
        return [], f"Excel读取失败：{e}"

def copy_project(project_dir, source_dirs, single_bar):
    project_name = os.path.basename(project_dir)
    excel_file = None
    for f in os.listdir(project_dir):
        if f.endswith(".xlsx"):
            excel_file = os.path.join(project_dir, f)
            break
    if not excel_file:
        log_and_display(f"[失败] {project_name} - 未找到 Excel 文件")
        return

    identifiers, err = extract_identifiers_from_excel(excel_file)
    if err:
        log_and_display(f"[失败] {project_name} - {err}")
        return

    dest_folder = os.path.join(project_dir, "PZIMG")
    os.makedirs(dest_folder, exist_ok=True)

    single_bar["maximum"] = len(identifiers)
    single_bar["value"] = 0
    root.update_idletasks()

    for idx, identifier in enumerate(identifiers, 1):
        found = False
        for src_base in source_dirs:
            src_path = os.path.join(src_base, identifier)
            if os.path.isdir(src_path):
                dst_path = os.path.join(dest_folder, identifier)
                if os.path.exists(dst_path):
                    log_and_display(f"[跳过] {project_name} - {identifier} 已存在，跳过复制")
                else:
                    try:
                        shutil.copytree(src_path, dst_path)
                        log_and_display(f"[成功] {project_name} - {identifier} ← {src_path}")
                    except Exception as e:
                        log_and_display(f"[错误] {project_name} - 复制失败：{e}")
                found = True
                break
        if not found:
            log_and_display(f"[失败] {project_name} - {identifier} 未在任一源目录中找到")
        single_bar["value"] = idx
        single_bar_label["text"] = f"{int((idx / len(identifiers)) * 100)}%"
        root.update_idletasks()

def log_and_display(message):
    log_entries.append(message)
    log_output.insert(tk.END, message + "\n")
    log_output.see(tk.END)

def run_copy():
    log_entries.clear()
    log_output.delete(1.0, tk.END)

    if not project_paths or not source_paths:
        messagebox.showwarning("缺少路径", "请先选择教师文件目录和财务文件目录。")
        return

    total = 0
    for root_dir in project_paths:
        for sub in os.listdir(root_dir):
            if os.path.isdir(os.path.join(root_dir, sub)):
                total += 1
    total_bar["maximum"] = total
    total_bar["value"] = 0
    root.update_idletasks()

    done = 0
    for root_dir in project_paths:
        for sub in os.listdir(root_dir):
            sub_path = os.path.join(root_dir, sub)
            if os.path.isdir(sub_path):
                copy_project(sub_path, source_paths, single_bar)
                done += 1
                total_bar["value"] = done
                total_bar_label["text"] = f"{int((done / total) * 100)}%"
                root.update_idletasks()

def save_log():
    if not log_entries:
        messagebox.showinfo("无日志", "当前没有日志可保存。")
        return
    default_name = f"日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    file_path = filedialog.asksaveasfilename(
        title="保存日志文件",
        initialdir=os.getcwd(),  # ✅ 修改点
        initialfile=default_name,
        defaultextension=".txt",
        filetypes=[("Text files", "*.txt")]
    )

    if file_path:
        with open(file_path, "w", encoding="utf-8") as f:
            for line in log_entries:
                f.write(line + "\n")
        messagebox.showinfo("保存成功", f"日志已保存到 {file_path}")

# ---------------- UI ---------------- #
root = tk.Tk()
root.title("📁 财务文件复制工具")
root.geometry("960x720")

main_frame = tk.Frame(root)
main_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

top_frame = tk.Frame(main_frame)
top_frame.pack(fill=tk.X)

# -------- 项目路径 --------
project_frame = tk.Frame(top_frame)
project_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
tk.Label(project_frame, text="教师文件根目录").pack(anchor="w")

project_canvas = tk.Canvas(project_frame, height=200, bg="white", bd=1, relief="sunken")
project_scrollbar = ttk.Scrollbar(project_frame, orient="vertical", command=project_canvas.yview)
project_scrollable_frame = tk.Frame(project_canvas, bg="white")
project_window = project_canvas.create_window((0, 0), window=project_scrollable_frame, anchor="nw")

project_canvas.configure(yscrollcommand=project_scrollbar.set)
project_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
project_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
project_canvas.bind("<Configure>", lambda e: project_canvas.itemconfig(project_window, width=e.width))
project_scrollable_frame.bind("<Configure>", lambda e: project_canvas.configure(scrollregion=project_canvas.bbox("all")))

# 添加/清空按钮在白框外部底部，水平排列
project_btn_frame = tk.Frame(project_frame)
project_btn_frame.pack(side="top", fill="x", pady=5)
tk.Button(project_btn_frame, text="添加", command=select_project_dirs).pack(side=tk.LEFT, padx=2)
tk.Button(project_btn_frame, text="清空", command=clear_project_dirs).pack(side=tk.LEFT, padx=2)

# -------- 源路径 --------
source_frame = tk.Frame(top_frame)
source_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
tk.Label(source_frame, text="财务文件根目录").pack(anchor="w")

source_canvas = tk.Canvas(source_frame, height=200, bg="white", bd=1, relief="sunken")
source_scrollbar = ttk.Scrollbar(source_frame, orient="vertical", command=source_canvas.yview)
source_scrollable_frame = tk.Frame(source_canvas, bg="white")
source_window = source_canvas.create_window((0, 0), window=source_scrollable_frame, anchor="nw")

source_canvas.configure(yscrollcommand=source_scrollbar.set)
source_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
source_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
source_canvas.bind("<Configure>", lambda e: source_canvas.itemconfig(source_window, width=e.width))
source_scrollable_frame.bind("<Configure>", lambda e: source_canvas.configure(scrollregion=source_canvas.bbox("all")))

# 添加/清空按钮在白框外部底部，水平排列
source_btn_frame = tk.Frame(source_frame)
source_btn_frame.pack(side="top", fill="x", pady=5)
tk.Button(source_btn_frame, text="添加", command=select_source_dirs).pack(side=tk.LEFT, padx=2)
tk.Button(source_btn_frame, text="清空", command=clear_source_dirs).pack(side=tk.LEFT, padx=2)

# -------- 控制进度 --------
middle_frame = tk.Frame(main_frame)
middle_frame.pack(fill=tk.X, pady=10)
tk.Button(middle_frame, text="🚀 开始复制", bg="green", fg="white", command=run_copy).pack(pady=5)

tk.Label(middle_frame, text="总任务进度").pack()
total_bar = ttk.Progressbar(middle_frame, length=900, mode='determinate')
total_bar.pack(pady=2)
total_bar_label = tk.Label(middle_frame, text="0%")
total_bar_label.pack()

tk.Label(middle_frame, text="当前任务进度").pack()
single_bar = ttk.Progressbar(middle_frame, length=900, mode='determinate')
single_bar.pack(pady=2)
single_bar_label = tk.Label(middle_frame, text="0%")
single_bar_label.pack()

# -------- 日志区 --------
bottom_frame = tk.Frame(main_frame)
bottom_frame.pack(fill=tk.BOTH, expand=True)

log_button_frame = tk.Frame(bottom_frame)
log_button_frame.pack(fill=tk.X)
tk.Button(log_button_frame, text="💾 下载日志", command=save_log).pack(side=tk.RIGHT, padx=5, pady=5)

log_output = scrolledtext.ScrolledText(bottom_frame, width=120, height=20)
log_output.pack(fill=tk.BOTH, expand=True)

root.mainloop()
